import datetime
import io
from typing import List

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from .. import models, schemas, security
from ..database import get_db
from ..services import mqtt_service

router = APIRouter(prefix="/api/doors", tags=["doors"])


@router.get("", response_model=List[schemas.DoorOut])
def list_doors(db: Session = Depends(get_db), user=Depends(security.get_current_user)):
    query = db.query(models.Door)
    if user.role in ("instructor", "doctor"):
        # Instructors and doctors only see doors an admin has explicitly
        # assigned them — not the full building. See models.DoorAssignment.
        query = query.join(
            models.DoorAssignment, models.DoorAssignment.door_id == models.Door.door_id
        ).filter(models.DoorAssignment.instructor_id == user.user_id)
    return query.all()


@router.post("", response_model=schemas.DoorOut, status_code=201)
def create_door(payload: schemas.DoorCreate, db: Session = Depends(get_db),
                 admin=Depends(security.require_admin)):
    if payload.fail_mode not in ("secure", "safe"):
        raise HTTPException(status_code=400, detail="fail_mode must be 'secure' or 'safe'")
    if payload.category not in ("critical", "access_service"):
        raise HTTPException(status_code=400, detail="category must be 'critical' or 'access_service'")

    existing = db.query(models.Door).filter(models.Door.code == payload.code).first()
    if existing:
        raise HTTPException(status_code=409, detail=f"A door with code '{payload.code}' already exists")

    door = models.Door(
        code=payload.code, name=payload.name, building=payload.building, floor=payload.floor,
        fail_mode=payload.fail_mode, category=payload.category, online=False, locked=True,
    )
    db.add(door)
    db.commit()
    db.refresh(door)
    return door


_CODE_ALIASES = {"code", "door code", "doorcode", "id", "door id"}
_NAME_ALIASES = {"name", "door name", "room", "room name", "room number"}
_BUILDING_ALIASES = {"building", "location"}
_FLOOR_ALIASES = {"floor", "level"}
_CATEGORY_ALIASES = {"category", "type", "classification"}
_FAIL_MODE_ALIASES = {"fail_mode", "fail mode", "failmode"}
_CRITICAL_HINTS = ("main", "entrance", "server", "critical", "gate")


def _classify(name_text: str, category_text: str | None) -> str:
    cat = (category_text or "").strip().lower().replace("-", "_").replace(" ", "_")
    if cat == "critical":
        return "critical"
    if cat in ("access_service", "access"):
        return "access_service"
    # No usable category cell — guess from the room name/building itself
    # ("Main Entrance", "Server Room" read as critical; everything else,
    # e.g. halls/lectures/sections, defaults to access_service).
    combined = f"{name_text} {category_text or ''}".lower()
    if any(hint in combined for hint in _CRITICAL_HINTS):
        return "critical"
    return "access_service"


@router.post("/import")
async def import_doors(file: UploadFile = File(...), db: Session = Depends(get_db),
                        admin=Depends(security.require_admin)):
    """Bulk-creates doors from an uploaded .xlsx sheet — for universities
    with far too many doors/halls/section rooms to add one at a time by
    hand. Expected columns (header row, any order, case-insensitive):
    Code, Name, Building, and optionally Category (values "critical" or
    "access_service" — anything else, or a blank cell, gets guessed from
    the room name instead) and Fail_mode ("secure"/"safe", default secure).
    """
    if not (file.filename or "").lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Please upload an .xlsx (Excel) file")

    contents = await file.read()
    try:
        wb = load_workbook(io.BytesIO(contents), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Couldn't read that file — is it a valid .xlsx?")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="That sheet is empty")

    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]

    def find_col(aliases):
        for i, h in enumerate(header):
            if h in aliases:
                return i
        return None

    col_code = find_col(_CODE_ALIASES)
    col_name = find_col(_NAME_ALIASES)
    col_building = find_col(_BUILDING_ALIASES)
    col_floor = find_col(_FLOOR_ALIASES)
    col_category = find_col(_CATEGORY_ALIASES)
    col_fail = find_col(_FAIL_MODE_ALIASES)

    if col_code is None or col_name is None or col_building is None:
        raise HTTPException(
            status_code=400,
            detail="Sheet needs columns for Code, Name, and Building "
                   "(Category is optional — guessed from the name if missing).",
        )

    def cell(row, idx):
        if idx is None or idx >= len(row):
            return None
        v = row[idx]
        return str(v).strip() if v is not None else None

    existing_codes = {c for (c,) in db.query(models.Door.code).all()}
    existing_buildings = {b for (b,) in db.query(models.Building.name).all()}

    created, skipped, errors = [], [], []
    for row_num, row in enumerate(rows[1:], start=2):
        code = cell(row, col_code)
        name = cell(row, col_name)
        building = cell(row, col_building)
        floor = cell(row, col_floor)
        category_raw = cell(row, col_category)
        fail_mode = (cell(row, col_fail) or "secure").lower()
        if fail_mode not in ("secure", "safe"):
            fail_mode = "secure"

        if not code or not name or not building:
            errors.append(f"Row {row_num}: missing code, name, or building — skipped")
            continue

        code = code.upper()
        if code in existing_codes:
            skipped.append(f"Row {row_num}: door code '{code}' already exists — skipped")
            continue

        door = models.Door(
            code=code, name=name, building=building, floor=floor, fail_mode=fail_mode,
            category=_classify(name, category_raw), online=False, locked=True,
        )
        db.add(door)
        existing_codes.add(code)
        created.append(code)

        if building not in existing_buildings:
            db.add(models.Building(name=building))
            existing_buildings.add(building)

    db.commit()
    return {"created": len(created), "created_codes": created, "skipped": skipped, "errors": errors}


@router.delete("/{door_id}", status_code=204)
def delete_door(door_id: int, db: Session = Depends(get_db), _admin=Depends(security.require_admin)):
    """Removes a door entirely. Drops door_assignments and alerts tied to it
    (meaningless once the door is gone), but leaves access_events alone —
    that's the audit trail and should survive even if the door itself is
    later removed, same spirit as unlinking (not deleting) credentials when
    a user is removed.
    """
    door = db.query(models.Door).filter(models.Door.door_id == door_id).first()
    if not door:
        raise HTTPException(status_code=404, detail="Door not found")

    db.query(models.DoorAssignment).filter(models.DoorAssignment.door_id == door_id).delete()
    db.query(models.Alert).filter(models.Alert.door_id == door_id).delete()
    db.query(models.Schedule).filter(models.Schedule.door_id == door_id).delete()
    db.delete(door)
    db.commit()


@router.get("/{door_id}", response_model=schemas.DoorOut)
def get_door(door_id: int, db: Session = Depends(get_db), _user=Depends(security.get_current_user)):
    door = db.query(models.Door).filter(models.Door.door_id == door_id).first()
    if not door:
        raise HTTPException(status_code=404, detail="Door not found")
    return door


@router.get("/{door_id}/logs", response_model=List[schemas.AccessEventOut])
def door_logs(door_id: int, db: Session = Depends(get_db), _user=Depends(security.get_current_user)):
    return (
        db.query(models.AccessEvent)
        .filter(models.AccessEvent.door_id == door_id)
        .order_by(models.AccessEvent.event_time.desc())
        .limit(200)
        .all()
    )


@router.post("/{door_id}/override")
def override_door(door_id: int, payload: schemas.DoorOverrideRequest, db: Session = Depends(get_db),
                   admin=Depends(security.require_admin)):
    door = db.query(models.Door).filter(models.Door.door_id == door_id).first()
    if not door:
        raise HTTPException(status_code=404, detail="Door not found")
    if payload.action not in ("lock", "unlock"):
        raise HTTPException(status_code=400, detail="action must be 'lock' or 'unlock'")

    sent = mqtt_service.publish_override(door.code, payload.action)

    # Log the override attempt regardless of delivery, so there's an audit
    # trail even if the broker/node didn't acknowledge it.
    event = models.AccessEvent(
        door_id=door.door_id,
        credential_id=None,
        method="override",
        result="sent" if sent else "queued_no_broker",
    )
    db.add(event)
    if payload.action == "unlock":
        door.locked = False
    else:
        door.locked = True
    db.commit()

    return {"door_id": door_id, "action": payload.action, "mqtt_delivered": sent}


@router.post("/{door_id}/status", response_model=schemas.DoorOut)
def set_door_status(door_id: int, payload: schemas.DoorStatusUpdate, db: Session = Depends(get_db),
                     admin=Depends(security.require_admin)):
    """Manual online/offline override for demo/testing use — e.g. forcing a
    door back "online" without waiting for a real MQTT status message, or
    marking one offline for maintenance. Note the Phase 7 staleness watchdog
    will flip it back to offline again if nothing real reports in within
    DOOR_STALE_AFTER_SECONDS, since this only sets the flag, it doesn't fake
    an ongoing heartbeat.
    """
    door = db.query(models.Door).filter(models.Door.door_id == door_id).first()
    if not door:
        raise HTTPException(status_code=404, detail="Door not found")
    door.online = payload.online
    if payload.online:
        door.last_seen = datetime.datetime.utcnow()
    db.commit()
    db.refresh(door)
    return door


@router.post("/{door_id}/request-access", response_model=schemas.AlertOut, status_code=201)
def request_access(door_id: int, db: Session = Depends(get_db),
                    user=Depends(security.get_current_user)):
    """An instructor asks an admin to open a door they're assigned to.

    Deliberately does NOT unlock anything itself — it only raises an alert
    (type=access_requested) that the existing AlertBanner already polls and
    renders for admins, who then use their own Unlock button to actually act
    on it. Keeps "ask" and "act" as two distinct, separately-audited steps.
    """
    door = db.query(models.Door).filter(models.Door.door_id == door_id).first()
    if not door:
        raise HTTPException(status_code=404, detail="Door not found")

    alert = models.Alert(door_id=door_id, type="access_requested", requested_by=user.user_id)
    db.add(alert)
    db.commit()
    db.refresh(alert)
    return alert
