import io
import os
import random
import re
import string
import uuid
from pathlib import Path
from typing import List

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from .. import models, schemas, security
from ..database import get_db

router = APIRouter(prefix="/api/users", tags=["users"])

# backend/media/avatars — see app/main.py for the StaticFiles mount that
# serves this directory at /media.
AVATAR_DIR = Path(__file__).resolve().parent.parent.parent / "media" / "avatars"
AVATAR_DIR.mkdir(parents=True, exist_ok=True)
ALLOWED_PHOTO_EXTS = {".png", ".jpg", ".jpeg", ".gif", ".webp"}
MAX_PHOTO_BYTES = 5 * 1024 * 1024

STAFF_EMAIL_DOMAIN = "aiu.is"


def _suggest_email(name: str) -> str:
    slug = re.sub(r"[^a-z0-9\s.]", "", name.strip().lower())
    parts = [p for p in slug.split() if p]
    base = ".".join(parts) or "user"
    return f"{base}@{STAFF_EMAIL_DOMAIN}"


def _random_password(length: int = 10) -> str:
    alphabet = string.ascii_letters + string.digits
    return "".join(random.choice(alphabet) for _ in range(length))


@router.get("", response_model=List[schemas.UserOut])
def list_users(db: Session = Depends(get_db), _admin=Depends(security.require_admin)):
    return db.query(models.User).all()


# ---------- "Me" (the logged-in user managing their own account) ----------
# Declared before "/{user_id}/..." routes so "/me" isn't ever swallowed by a
# path-param route.

@router.get("/me", response_model=schemas.UserOut)
def get_me(user=Depends(security.get_current_user)):
    return user


@router.patch("/me/profile", response_model=schemas.UserOut)
def update_my_profile(payload: schemas.ProfileUpdate, db: Session = Depends(get_db),
                       user=Depends(security.get_current_user)):
    """Updates name/email for the logged-in user. Note the JWT's subject is
    the email, so changing it invalidates the *meaning* of the old token
    (it'll still decode, but won't match any user by email) — the frontend
    must call POST /api/auth/refresh right after this to get a token that
    matches the new email, or the next request will 401.
    """
    if payload.name is not None:
        name = payload.name.strip()
        if not name:
            raise HTTPException(status_code=400, detail="Name can't be empty")
        user.name = name

    if payload.email is not None and payload.email != user.email:
        existing = db.query(models.User).filter(
            models.User.email == payload.email, models.User.user_id != user.user_id
        ).first()
        if existing:
            raise HTTPException(status_code=409, detail=f"A user with email '{payload.email}' already exists")
        user.email = payload.email

    db.commit()
    db.refresh(user)
    return user


@router.patch("/me/password", status_code=204)
def change_my_password(payload: schemas.PasswordChange, db: Session = Depends(get_db),
                        user=Depends(security.get_current_user)):
    if not security.verify_password(payload.current_password, user.password_hash):
        raise HTTPException(status_code=400, detail="Current password is incorrect")
    if len(payload.new_password) < 6:
        raise HTTPException(status_code=400, detail="New password must be at least 6 characters")
    user.password_hash = security.hash_password(payload.new_password)
    db.commit()


@router.post("/me/photo", response_model=schemas.UserOut)
async def upload_my_photo(file: UploadFile = File(...), db: Session = Depends(get_db),
                           user=Depends(security.get_current_user)):
    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext not in ALLOWED_PHOTO_EXTS:
        raise HTTPException(status_code=400, detail="Photo must be a PNG, JPG, GIF, or WEBP image")

    contents = await file.read()
    if len(contents) > MAX_PHOTO_BYTES:
        raise HTTPException(status_code=400, detail="Photo must be under 5MB")

    filename = f"{user.user_id}-{uuid.uuid4().hex}{ext}"
    with open(AVATAR_DIR / filename, "wb") as f:
        f.write(contents)

    user.photo_url = f"/media/avatars/{filename}"
    db.commit()
    db.refresh(user)
    return user


@router.post("", response_model=schemas.UserOut, status_code=201)
def create_user(payload: schemas.UserCreate, db: Session = Depends(get_db), _admin=Depends(security.require_admin)):
    if payload.role not in ("admin", "instructor", "doctor"):
        raise HTTPException(status_code=400, detail="role must be 'admin', 'instructor', or 'doctor'")
    existing = db.query(models.User).filter(models.User.email == payload.email).first()
    if existing:
        raise HTTPException(status_code=409, detail=f"A user with email '{payload.email}' already exists")
    user = models.User(
        name=payload.name,
        email=payload.email,
        role=payload.role,
        password_hash=security.hash_password(payload.password),
        faculty_id=payload.faculty_id,
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return user


@router.post("/import")
async def import_users(role: str = Form(...), file: UploadFile = File(...),
                        db: Session = Depends(get_db), _admin=Depends(security.require_admin)):
    """Bulk-creates TAs or doctors from an uploaded .xlsx — for importing a
    whole staff roster at once instead of adding each one by hand. Expected
    columns (header row, any order, case-insensitive): Name (required), and
    optionally Email (auto-generated as name@aiu.is if blank), Password
    (a random one is generated if blank), and Faculty (created if it
    doesn't already exist). Generated passwords are returned in the
    response since they can't be recovered later — bcrypt only stores the
    hash — so the admin needs to copy them out immediately to hand out.
    """
    if role not in ("instructor", "doctor"):
        raise HTTPException(status_code=400, detail="role must be 'instructor' or 'doctor'")
    if not (file.filename or "").lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Please upload an .xlsx (Excel) file")

    contents = await file.read()
    try:
        wb = load_workbook(io.BytesIO(contents), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Couldn't read that file — is it a valid .xlsx?")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="That sheet is empty")

    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]

    def find_col(aliases):
        for i, h in enumerate(header):
            if h in aliases:
                return i
        return None

    col_name = find_col({"name", "full name", "fullname"})
    col_email = find_col({"email"})
    col_password = find_col({"password"})
    col_faculty = find_col({"faculty"})

    if col_name is None:
        raise HTTPException(
            status_code=400,
            detail="Sheet needs a Name column (Email, Password, and Faculty are optional).",
        )

    def cell(row, idx):
        if idx is None or idx >= len(row):
            return None
        v = row[idx]
        return str(v).strip() if v is not None else None

    existing_emails = {e for (e,) in db.query(models.User.email).all()}
    faculty_by_name = {f.name: f for f in db.query(models.Faculty).all()}

    created, skipped, errors = [], [], []
    for row_num, row in enumerate(rows[1:], start=2):
        name = cell(row, col_name)
        if not name:
            errors.append(f"Row {row_num}: missing name — skipped")
            continue

        email = (cell(row, col_email) or _suggest_email(name)).lower()
        if email in existing_emails:
            skipped.append(f"Row {row_num}: email '{email}' already exists — skipped")
            continue

        password = cell(row, col_password) or _random_password()

        faculty_name = cell(row, col_faculty)
        faculty_id = None
        if faculty_name:
            faculty = faculty_by_name.get(faculty_name)
            if not faculty:
                faculty = models.Faculty(name=faculty_name)
                db.add(faculty)
                db.flush()
                faculty_by_name[faculty_name] = faculty
            faculty_id = faculty.faculty_id

        user = models.User(
            name=name, email=email, role=role,
            password_hash=security.hash_password(password),
            faculty_id=faculty_id,
        )
        db.add(user)
        existing_emails.add(email)
        created.append({"name": name, "email": email, "password": password})

    db.commit()
    return {"created": created, "skipped": skipped, "errors": errors}


# ---------- Door assignments (which doors a TA is allowed to request) ----------

@router.get("/{user_id}/doors", response_model=List[schemas.DoorAssignmentOut])
def list_door_assignments(user_id: int, db: Session = Depends(get_db),
                           _admin=Depends(security.require_admin)):
    return (
        db.query(models.DoorAssignment)
        .filter(models.DoorAssignment.instructor_id == user_id)
        .all()
    )


@router.post("/{user_id}/doors", response_model=schemas.DoorAssignmentOut, status_code=201)
def add_door_assignment(user_id: int, payload: schemas.DoorAssignmentCreate,
                         db: Session = Depends(get_db), _admin=Depends(security.require_admin)):
    target = db.query(models.User).filter(models.User.user_id == user_id).first()
    if not target:
        raise HTTPException(status_code=404, detail="User not found")
    door = db.query(models.Door).filter(models.Door.door_id == payload.door_id).first()
    if not door:
        raise HTTPException(status_code=404, detail="Door not found")

    existing = (
        db.query(models.DoorAssignment)
        .filter(
            models.DoorAssignment.instructor_id == user_id,
            models.DoorAssignment.door_id == payload.door_id,
        )
        .first()
    )
    if existing:
        raise HTTPException(status_code=409, detail="This door is already assigned to this user")

    assignment = models.DoorAssignment(instructor_id=user_id, door_id=payload.door_id)
    db.add(assignment)
    db.commit()
    db.refresh(assignment)
    return assignment


@router.delete("/{user_id}", status_code=204)
def delete_user(user_id: int, db: Session = Depends(get_db), admin=Depends(security.require_admin)):
    """Removes a TA/doctor (or any user) entirely — not just a door
    assignment. Unlinks their credentials (kept for the access_events audit
    trail, just no longer tied to a person) and drops their door
    assignments, since those are meaningless once the person is gone.
    """
    if user_id == admin.user_id:
        raise HTTPException(status_code=400, detail="You can't delete your own account while logged in as it")

    target = db.query(models.User).filter(models.User.user_id == user_id).first()
    if not target:
        raise HTTPException(status_code=404, detail="User not found")

    if target.role == "admin":
        remaining_admins = db.query(models.User).filter(
            models.User.role == "admin", models.User.user_id != user_id
        ).count()
        if remaining_admins == 0:
            raise HTTPException(status_code=400, detail="Can't delete the last remaining admin account")

    db.query(models.DoorAssignment).filter(models.DoorAssignment.instructor_id == user_id).delete()
    db.query(models.Credential).filter(models.Credential.user_id == user_id).update({"user_id": None})
    db.delete(target)
    db.commit()


@router.delete("/{user_id}/doors/{assignment_id}", status_code=204)
def remove_door_assignment(user_id: int, assignment_id: int, db: Session = Depends(get_db),
                            _admin=Depends(security.require_admin)):
    assignment = (
        db.query(models.DoorAssignment)
        .filter(
            models.DoorAssignment.assignment_id == assignment_id,
            models.DoorAssignment.instructor_id == user_id,
        )
        .first()
    )
    if not assignment:
        raise HTTPException(status_code=404, detail="Assignment not found")
    db.delete(assignment)
    db.commit()
